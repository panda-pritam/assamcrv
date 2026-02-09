
from django.http import HttpResponseForbidden
from accounts.models import tblUser

from django.utils.translation import get_language

def translated(obj, base_field):
    lang = get_language()
    return getattr(obj, f"{base_field}_{lang}", None) or getattr(obj, base_field, None)


def is_admin_or_superuser(user):
    return user.is_authenticated and (user.is_superuser or user.is_staff or user.role.name == "ASDMA")

def apply_location_filters(queryset, district_id=None, circle_id=None, gram_panchayat_id=None, village_id=None):
    if district_id:
        queryset = queryset.filter(village__gram_panchayat__circle__district_id=district_id)
    if circle_id:
        queryset = queryset.filter(village__gram_panchayat__circle_id=circle_id)
    if gram_panchayat_id:
        queryset = queryset.filter(village__gram_panchayat_id=gram_panchayat_id)
    if village_id:
        queryset = queryset.filter(village_id=village_id)
    return queryset

def village_apply_location_filters(queryset, district_id=None, circle_id=None, gram_panchayat_id=None, village_id=None):

    if district_id:
        queryset = queryset.filter(gram_panchayat__circle__district_id= district_id)
    if circle_id:
        queryset = queryset.filter(gram_panchayat__circle_id= circle_id)
    if gram_panchayat_id:
        queryset = queryset.filter(gram_panchayat_id= gram_panchayat_id)
    if village_id:        
        queryset = queryset.filter(id=village_id)
    return queryset


def apply_role_filters(user, role, queryset):
    if user.is_superuser or role == "ASDMA":
        return queryset
    elif role == "DDMA":
        return queryset.filter(village__gram_panchayat__circle__district=user.district)
    elif role == "Circle Officer":
        return queryset.filter(village__gram_panchayat__circle=user.circle)
    elif role == "Gram Panchayat Officer":
        return queryset.filter(village__gram_panchayat=user.gram_panchayat)
    elif role == "Village Officer":
        return queryset.filter(village=user.village)
    else:
        return queryset.none()

from typing import List, Optional

def get_village_codes(
    district_id: Optional[int] = None, 
    circle_id: Optional[int] = None, 
    gram_panchayat_id: Optional[int] = None, 
    village_id: Optional[int] = None
) -> List[str]:
    """
    Returns village codes based on the hierarchy:
    district > circle > gram_panchayat > village
    """
   

    if village_id:
        villages = tblVillage.objects.filter(id=village_id)
    elif gram_panchayat_id:
        villages = tblVillage.objects.filter(gram_panchayat_id=gram_panchayat_id)
    elif circle_id:
        villages = tblVillage.objects.filter(gram_panchayat__circle_id=circle_id)
    elif district_id:
        villages = tblVillage.objects.filter(gram_panchayat__circle__district_id=district_id)
    else:
        villages = tblVillage.objects.all()  

    return list(villages.values_list('code', flat=True))



def get_filtered_users(user, district_id=None, circle_id=None, gram_panchayat_id=None, village_id=None, role_id=None, department_id=None):
    """
    Returns filtered users based on the user's role and optional filter parameters.
    Used to retrieve users accessible to the current user with hierarchical role checks.
    """
    users = tblUser.objects.none()  # Default empty queryset

    if user.is_authenticated:
        role = getattr(user.role, 'name', None)
        if user.is_superuser or role == "ASDMA":
            users = tblUser.objects.all()
        elif role == "DDMA":
            users = tblUser.objects.filter(district_id=user.district_id)
        elif role == "Circle Officer":
            users = tblUser.objects.filter(circle_id=user.circle_id)
        elif role == "Gram Panchayat Officer":
            users = tblUser.objects.filter(gram_panchayat_id=user.gram_panchayat_id)
        elif role == "Village Officer":
            users = tblUser.objects.filter(village_id=user.village_id)

    if district_id:
        users = users.filter(district_id=district_id)
    if circle_id:
        users = users.filter(circle_id=circle_id)
    if gram_panchayat_id:
        users = users.filter(gram_panchayat_id=gram_panchayat_id)
    if village_id:
        users = users.filter(village_id=village_id)
    if role_id:
        users = users.filter(role_id=role_id)
    if department_id:
        users = users.filter(department_id=department_id)

    return users


import csv
import os
import tempfile
from django.db import transaction
from googletrans import Translator
from accounts.models import tblDistrict, tblCircle, tblGramPanchayat, tblVillage

from openpyxl import load_workbook
import logging

# Set up logging
logger = logging.getLogger(__name__)

translator = Translator()

def translate_name(name, lang_code):
    """Translate name to given language code with error handling"""
    # Temporarily disable translation due to coroutine issues
    # Return original name as fallback
    return name.strip() if name else name

def clean_cell_value(cell_value):
    """Clean and convert cell value to string"""
    if cell_value is None:
        return ""
    if isinstance(cell_value, (int, float)):
        # Convert numbers to string, handle floats properly
        if isinstance(cell_value, float) and cell_value.is_integer():
            return str(int(cell_value))
        return str(cell_value)
    return str(cell_value).strip()


def read_csv_file(file_path):
    """Read CSV file with multiple encoding attempts"""
    encodings = ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
    
    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding, newline='') as csvfile:
                # Test if we can read the file
                sample = csvfile.read(1024)
                csvfile.seek(0)
                
                # Create CSV reader
                reader = csv.DictReader(csvfile)
                rows = []
                
                # Read all rows
                for row_num, row in enumerate(reader, start=1):
                    if not any(value.strip() for value in row.values() if value):
                        continue  # Skip empty rows
                    
                    # Clean row data
                    cleaned_row = {}
                    for key, value in row.items():
                        cleaned_row[key] = clean_cell_value(value)
                    
                    rows.append(cleaned_row)
                
                logger.info(f"Successfully read {len(rows)} rows from CSV file using {encoding} encoding")
                return rows
                
        except UnicodeDecodeError:
            continue
        except Exception as e:
            logger.error(f"Error reading CSV with {encoding}: {e}")
            continue
    
    raise ValueError(f"Could not decode CSV file with any of the supported encodings: {encodings}")

def convert_excel_to_csv(excel_path, csv_path):
    """Convert Excel file to CSV format for reliable processing"""
    try:
        logger.info(f"Converting Excel file {excel_path} to CSV format...")
        
        # Load the workbook
        workbook = load_workbook(excel_path, data_only=True, read_only=True)
        
        # Get the active sheet
        if workbook.sheetnames:
            sheet = workbook.active
            logger.info(f"Using sheet: {sheet.title}")
        else:
            raise ValueError("Excel file has no sheets")
        
        # Write to CSV
        with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
            csv_writer = csv.writer(csvfile)
            
            row_count = 0
            for row in sheet.iter_rows(values_only=True):
                # Skip completely empty rows
                if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                    continue
                
                # Clean and convert row data
                cleaned_row = []
                for cell in row:
                    cleaned_value = clean_cell_value(cell)
                    cleaned_row.append(cleaned_value)
                
                csv_writer.writerow(cleaned_row)
                row_count += 1
        
        workbook.close()
        logger.info(f"Successfully converted Excel to CSV: {row_count} rows written")
        return True
        
    except Exception as e:
        logger.error(f"Error converting Excel to CSV: {e}")
        raise ValueError(f"Error converting Excel to CSV: {str(e)}")

def import_location_data(file_path, update_existing=True):
    """Import location data from CSV or Excel file"""
    
    # Validate file exists
    if not os.path.exists(file_path):
        raise ValueError(f"File does not exist: {file_path}")
    
    # Get file extension
    _, ext = os.path.splitext(file_path.lower())
    
    # Convert Excel files to CSV first for more reliable processing
    csv_file_path = file_path
    temp_csv_path = None
    
    try:
        if ext in ['.xlsx', '.xls']:
            logger.info(f"Excel file detected. Converting to CSV for processing...")
            
            # Create temporary CSV file
            temp_csv_fd, temp_csv_path = tempfile.mkstemp(suffix='.csv', prefix='converted_')
            os.close(temp_csv_fd)  # Close the file descriptor, we'll use the path
            
            # Convert Excel to CSV
            convert_excel_to_csv(file_path, temp_csv_path)
            csv_file_path = temp_csv_path
            
        elif ext == '.csv':
            logger.info("CSV file detected. Processing directly...")
        else:
            raise ValueError(f"Unsupported file format '{ext}'. Only .csv, .xlsx, and .xls files are supported.")
        
        # Now read the CSV file (either original or converted)
        rows = read_csv_file(csv_file_path)
        
        if not rows:
            raise ValueError("No data rows found in the file")
        
    except Exception as e:
        logger.error(f"File reading error: {e}")
        raise
    finally:
        # Clean up temporary CSV file if created
        if temp_csv_path and os.path.exists(temp_csv_path):
            try:
                os.remove(temp_csv_path)
                logger.info("Temporary CSV file cleaned up")
            except Exception as e:
                logger.warning(f"Failed to clean up temporary CSV file: {e}")
    
    # Process the data
    success_count = 0
    error_count = 0
    total_rows = len(rows)
    
    logger.info(f"Starting to process {total_rows} rows")
    
    def get_field_value(row, field_list):
        for field in field_list:
            if field in row and row[field]:
                return row[field].strip()
        return ""
    
    for row_num, row in enumerate(rows, start=1):
        try:
            # Extract data from new Excel format
            district_name = get_field_value(row, ['District_name', 'District', 'district_name']).title()
            district_name_bn = get_field_value(row, ['District_name_bn', 'district_name_bn'])
            district_name_as = get_field_value(row, ['District_name_as', 'district_name_as'])
            district_code = get_field_value(row, ['District_Code', 'district_code'])
            
            circle_name = get_field_value(row, ['Revenue_circle', 'Circle_Name', 'circle_name']).title()
            circle_name_bn = get_field_value(row, ['Revenue_circle_bn', 'circle_name_bn'])
            circle_name_as = get_field_value(row, ['Revenue_circle_as', 'circle_name_as'])
            
            village_name = get_field_value(row, ['Village_name', 'village_name']).title()
            village_name_bn = get_field_value(row, ['Village_name_bn', 'village_name_bn'])
            village_name_as = get_field_value(row, ['Village_name_as', 'village_name_as'])
            village_code = get_field_value(row, ['Village_Code', 'village_code'])
            
            gp_name = get_field_value(row, ['Block_Name', 'block_name']).title() or village_name
            
            # Parse coordinates
            def parse_coord(value):
                try:
                    return float(value) if value else None
                except (TypeError, ValueError):
                    return None

            district_lat = parse_coord(row.get("Dist_lat"))
            district_lng = parse_coord(row.get("Dist_lng"))
            village_lat = parse_coord(row.get("Vill_lat"))
            village_lng = parse_coord(row.get("Vill_lng"))
            
            # Use English name as fallback for translations
            district_name_bn = district_name_bn or district_name
            district_name_as = district_name_as or district_name
            circle_name_bn = circle_name_bn or circle_name
            circle_name_as = circle_name_as or circle_name
            village_name_bn = village_name_bn or village_name
            village_name_as = village_name_as or village_name
            gp_name_bn = village_name_bn
            gp_name_as = village_name_as
            
            # Validate required fields
            if not all([district_name, circle_name, village_name]):
                logger.warning(f"Row {row_num}: Missing required fields - District: '{district_name}', Circle: '{circle_name}', Village: '{village_name}'")
                logger.warning(f"Row data: {row}")
                error_count += 1
                continue
            
            # Process the row
            with transaction.atomic():
                # Create/update District
                district = None
                if district_code:
                    district = tblDistrict.objects.filter(code=district_code).first()
                
                if not district:
                    district, created = tblDistrict.objects.get_or_create(
                        name=district_name,
                        defaults={
                            'code': district_code,
                            'name_bn': district_name_bn,
                            'name_as': district_name_as,
                            'latitude': district_lat,
                            'longitude': district_lng
                        }
                    )
                    if not created and update_existing:
                        district.code = district_code or district.code
                        district.name_bn = district_name_bn
                        district.name_as = district_name_as
                        district.latitude = district_lat
                        district.longitude = district_lng
                        district.save()
                elif update_existing:
                    district.name = district_name
                    district.name_bn = district_name_bn
                    district.name_as = district_name_as
                    district.latitude = district_lat
                    district.longitude = district_lng
                    district.save()
                
                # Create/update Circle
                circle, created = tblCircle.objects.get_or_create(
                    name=circle_name,
                    district=district,
                    defaults={
                        'name_bn': circle_name_bn,
                        'name_as': circle_name_as
                    }
                )
                if not created and update_existing:
                    circle.name_bn = circle_name_bn
                    circle.name_as = circle_name_as
                    circle.save()
                
                # Create/update Gram Panchayat
                gp, created = tblGramPanchayat.objects.get_or_create(
                    name=gp_name,
                    circle=circle,
                    defaults={
                        'name_bn': gp_name_bn,
                        'name_as': gp_name_as
                    }
                )
                if not created and update_existing:
                    gp.name_bn = gp_name_bn
                    gp.name_as = gp_name_as
                    gp.save()
                
                # Create/update Village
                village = None
                if village_code:
                    village = tblVillage.objects.filter(code=village_code, gram_panchayat=gp).first()
                
                if not village:
                    village, created = tblVillage.objects.get_or_create(
                        name=village_name,
                        gram_panchayat=gp,
                        defaults={
                            'code': village_code,
                            'name_bn': village_name_bn,
                            'name_as': village_name_as,
                            'latitude': village_lat,
                            'longitude': village_lng
                        }
                    )
                    if not created and update_existing:
                        village.code = village_code or village.code
                        village.name_bn = village_name_bn
                        village.name_as = village_name_as
                        village.latitude = village_lat
                        village.longitude = village_lng
                        village.save()
                elif update_existing:
                    village.name = village_name
                    village.name_bn = village_name_bn
                    village.name_as = village_name_as
                    village.latitude = village_lat
                    village.longitude = village_lng
                    village.save()
                
                success_count += 1
                logger.info(f"Successfully processed row {row_num}: {district_name} -> {circle_name} -> {gp_name} -> {village_name}")
        
        except Exception as e:
            error_count += 1
            logger.error(f"Error processing row {row_num}: {e}")
            logger.error(f"Row data: {row}")
            continue  # Continue with next row
    
    # Final report
    logger.info(f"Import completed. Total: {total_rows}, Success: {success_count}, Errors: {error_count}")
    
    if error_count > 0:
        logger.warning(f"Warning: {error_count} rows had errors and were skipped.")
    
    if success_count == 0:
        raise ValueError("No rows were successfully imported. Please check your file format and data.")

HOUSEHOLD_MAPPING = {
    'district_code': 'district_code',
    'village_code': 'village_code',
    'point_id': 'point_id',
    'property_owner': 'property_owner',
    'name_of_hohh': 'name_of_hohh',
    'photo': 'photo',
    'mobile_number': 'mobile_number',
    'data_access': 'data_access',
    'community': 'community',
    'social_status': 'social_status',
    'economic_status': 'economic_status',
    'wall_type': 'wall_type',
    'roof_type': 'roof_type',
    'floor_type': 'floor_type',
    'plinth_or_stilt': 'plinth_or_stilt',
    'plinth_or_stilt_height_ft': 'plinth_or_stilt_height_ft',
    'number_of_storeys': 'number_of_storeys',
    'number_of_males_including_children': 'number_of_males_including_children',
    'number_of_females_including_children': 'number_of_females_including_children',
    'children_below_6_years': 'children_below_6_years',
    'senior_citizens': 'senior_citizens',
    'pregnant_women': 'pregnant_women',
    'lactating_women': 'lactating_women',
    'persons_with_disability_or_chronic_disease': 'persons_with_disability_or_chronic_disease',
    'drinking_water_source': 'drinking_water_source',
    'sanitation_facility': 'sanitation_facility',
    'toilet_wall_material': 'toilet_wall_material',
    'toilet_roof_material': 'toilet_roof_material',
    'digital_media_owned': 'digital_media_owned',
    'house_has_electric_connection': 'house_has_electric_connection',
    'source_of_electricity': 'source_of_electricity',
    'own_agriculture_land': 'own_agriculture_land',
    'area_of_agriculture_land_owned_bigha': 'area_of_agriculture_land_owned_bigha',
    'land_area_annually_cultivated_bigha': 'land_area_annually_cultivated_bigha',
    'crops_cultivated': 'crops_cultivated',
    'specify_other': 'specify_other',
    'number_of_crops_normally_raised_every_year': 'number_of_crops_normally_raised_every_year',
    'livelihood_primary': 'livelihood_primary',
    'livelihood_secondary': 'livelihood_secondary',
    'do_you_have_big_cattle_cattle_buffalo': 'do_you_have_big_cattle_cattle_buffalo',
    'number_of_big_cattle_animals': 'number_of_big_cattle_animals',
    'do_you_have_small_cattle_goat_sheep_pig': 'do_you_have_small_cattle_goat_sheep_pig',
    'number_of_small_cattle_animals': 'number_of_small_cattle_animals',
    'do_you_have_poultry_chicken_and_duck': 'do_you_have_poultry_chicken_and_duck',
    'number_of_poultry_animals': 'number_of_poultry_animals',
    'approximate_income_earned_every_year_inr': 'approximate_income_earned_every_year_inr',
    'expense_on_education': 'expense_on_education',
    'expense_on_health': 'expense_on_health',
    'expense_on_food': 'expense_on_food',
    'expense_on_tobacco_liquor': 'expense_on_tobacco_liquor',
    'expense_on_house_repair': 'expense_on_house_repair',
    'expense_on_festival_marriage_and_other_social_occassions': 'expense_on_festival_marriage_and_other_social_occassions',
    'amount_spent_for_agriculture_livestock': 'amount_spent_for_agriculture_livestock',
    'loss_due_to_flood': 'loss_due_to_flood',
    'loan_availed': 'loan_availed',
    'loan_amount': 'loan_amount',
    'loan_purpose': 'loan_purpose',
    'house_affected_by_flood': 'house_affected_by_flood',
    'economic_loss_to_your_house_due_to_flood': 'economic_loss_to_your_house_due_to_flood',
    'amount_towards_flood_recovery_expenditure': 'amount_towards_flood_recovery_expenditure',
    'maximum_flood_height_in_house_ft': 'maximum_flood_height_in_house_ft',
    'year_in_which_maximum_flood_experience_in_your_house': 'year_in_which_maximum_flood_experience_in_your_house',
    'your_agriculture_affected_by_flood': 'your_agriculture_affected_by_flood',
    'maximum_flood_height_experience_in_your_agriculture_ft': 'maximum_flood_height_experience_in_your_agriculture_ft',
    'year_in_which_max_flood_experience_in_your_agriculture_land': 'year_in_which_max_flood_experience_in_your_agriculture_land',
    'duration_of_flood_stay_in_your_agriculture_field': 'duration_of_flood_stay_in_your_agriculture_field',
    'other_natural_hazards_directly_impacting_you_and_family': 'other_natural_hazards_directly_impacting_you_and_family',
    'house_vulnerable_to_erosion': 'house_vulnerable_to_erosion',
    'your_agriculture_field_vulnerable_to_erosion': 'your_agriculture_field_vulnerable_to_erosion',
    'building_quality': 'building_quality',
    'foundation_quality': 'foundation_quality',
    'number_of_small_buildings_of_the_household': 'number_of_small_buildings_of_the_household',
    'occupa_ncy_type_of_small_building': 'occupa_ncy_type_of_small_building',
    'presence_of_grain_bank': 'presence_of_grain_bank',
    'plinth_height_of_grain_bank_ft': 'plinth_height_of_grain_bank_ft',
    'wall_material_of_grain_bank': 'wall_material_of_grain_bank',
    'roof_material_of_grain_bank': 'roof_material_of_grain_bank',
    'flood_depth_m': 'flood_depth_m',
    'flood_class': 'flood_class',
    'erosion_class': 'erosion_class',
    'loan_class': 'loan_class',
    'agrculture_land_class': 'agrculture_land_class',
    'fld_hh_class': 'fld_hh_class',
    'repair_class': 'repair_class',
    'economic_loss_hh': 'economic_loss_hh',
    'loss_agricultire_livlihood': 'loss_agricultire_livlihood',
    'big_cattle': 'big_cattle',
    'small_cattle': 'small_cattle',
    'house_type': 'house_type',
    'income_class': 'income_class',
    'crops_diversity': 'crops_diversity',
    'Sanitation_Type': 'Sanitation_Type',
    'form_id': 'form_id',
    'unique_id': 'unique_id',
    'latitude': 'latitude',
    'longitude': 'longitude',
    'build_area_meter': 'build_area_meter',
    'building_length_meter': 'building_length_meter',
    'building_width_meter': 'building_width_meter',
    'building_length_feet': 'building_length_feet',
    'building_width_feet': 'building_width_feet',
    'building_area_sqft': 'building_area_sqft',
    'flood_depth_from_survey_meter': 'flood_depth_from_survey_meter',
    'maximum_flood_height_meter': 'maximum_flood_height_meter',
    'plinth_or_stilt_height_meter': 'plinth_or_stilt_height_meter',
    'loss_AgriLivli': 'loss_AgriLivli',
    'toilet_class': 'toilet_class',
    'adequate_water_supply': 'adequate_water_supply',
    'JJM_or_other_taped_water_connection': 'JJM_or_other_taped_water_connection',
    'sludge_be_disposed_type': 'sludge_be_disposed_type',
    'type_of_toilet': 'type_of_toilet',
    # temp
    "You get adequate water from this source throughout the year": "adequate_water_supply",
    "Type of toilet?": "Sanitation_Type",
    "While de-sludging where the sludge be disposed -": "sludge_be_disposed_type",
}

COMMERCIAL_MAPPING = {
    'district_code': 'district_code',
    'village_code': 'village_code',
    'district_name': 'district_name',
    'village_name': 'village_name',
    'point_id': 'point_id',
    'type_of_occupancy': 'type_of_occupancy',
    'type_of_occupancy_others': 'type_of_occupancy_others',
    'property_owner': 'property_owner',
    'name_of_person': 'name_of_person',
    'photo': 'photo',
    'name_of_the_building': 'name_of_the_building',
    'name_of_the_in_charge': 'name_of_the_in_charge',
    'phone_number_of_the_in_charge': 'phone_number_of_the_in_charge',
    'wall_type': 'wall_type',
    'floor_type': 'floor_type',
    'roof_type': 'roof_type',
    'plinth_above_ground': 'plinth_above_ground',
    'plinth_above_ground_stilt_height_in_ft': 'plinth_above_ground_stilt_height_in_ft',
    'building_affected_by_normal_flood': 'building_affected_by_normal_flood',
    'approximate_content_value_inr': 'approximate_content_value_inr',
    'approximate_value_business_per_year': 'approximate_value_business_per_year',
    'average_room_width_ft': 'average_room_width_ft',
    'average_room_length_ft': 'average_room_length_ft',
    'building_quality': 'building_quality',
    'foundation_quality': 'foundation_quality',
    'access_road_during_flood': 'access_road_during_flood',
    'flood_depth_m': 'flood_depth_m',
    'erosion_class': 'erosion_class',
    'form_id': 'form_id',
    'unique_id': 'unique_id',
    'latitude': 'latitude',
    'longitude': 'longitude',
    'plinth_or_srilt_height_meter': 'plinth_or_srilt_height_meter',
    'erosion_value': 'erosion_value'
}

TRANSFORMER_MAPPING = {
    'Village_Name': 'village_name',
    'District_Name': 'district_name',
    'District_Code': 'district_code',
    'Village_Id': 'village_code',
    'Transformer_Site_Address': 'transformer_site_address',
    'Latitude': 'latitude',
    'Longitude': 'longitude',
    'Photo': 'photo',
    'Fencing': 'fencing',
    'Material': 'material',
    'Condition': 'condition',
    
}

CRITICAL_FACILITY = {
    'district_code': 'district_code',
    'district_name': 'district_name',
    'village_name': 'village_name',
    'village_code': 'village_code',
    'point_id': 'point_id',
    'occupancy_type': 'occupancy_type',
    'photo': 'photo',
    'name_of_building': 'name_of_building',
    'incharge_name': 'incharge_name',
    'mobile_number': 'mobile_number',
    'wall_type': 'wall_type',
    'floor_type': 'floor_type',
    'roof_type': 'roof_type',
    'plinth_or_stilt': 'plinth_or_stilt',
    'plinth_or_stilt_height_ft': 'plinth_or_stilt_height_ft',
    'drinking_water_source': 'drinking_water_source',
    'house_has_electric_connection': 'house_has_electric_connection',
    'source_of_electricity': 'source_of_electricity',
    'number_of_rooms': 'number_of_rooms',
    'average_room_length_ft': 'average_room_length_ft',
    'average_room_width_ft': 'average_room_width_ft',
    'kitchen_facility': 'kitchen_facility',
    'toilet_facility': 'toilet_facility',
    'number_of_toilets': 'number_of_toilets',
    'water_facility_in_toilet': 'water_facility_in_toilet',
    'electricity_facility_in_toilet': 'electricity_facility_in_toilet',
    'building_affected_by_normal_flood': 'building_affected_by_normal_flood',
    'used_as_a_flood_emergency_shelter': 'used_as_a_flood_emergency_shelter',
    'access_road_during_flood': 'access_road_during_flood',
    'building_quality': 'building_quality',
    'foundation_quality': 'foundation_quality',
    'flood_depth_m': 'flood_depth_m',
    'flood_class': 'flood_class',
    'erosion_class': 'erosion_class',
    'form_id': 'form_id',
    'unique_id': 'unique_id',
    'latitude': 'latitude',
    'longitude': 'longitude'
}

ELECTRIC_POLES = {
    'Village_Name': 'village_name',
    'District_Name': 'district_name',
    'District_Code': 'district_code',
    'Village_Id': 'village_code', 
    # 'Uid': 'uid',
    'Latitude': 'latitude',
    'Longitude': 'longitude',
    # 'Electric_Pole_Name': 'electric_pole_name',
    'Material': 'electric_pole_material',
    'Condition': 'electric_pole_condition',
    # 'Remarks_On_Pole_Condition': 'remarks_on_pole_condition',
    'Photo': 'photo',
    # 'Flood_Depth(m)': 'flood_depth_m',
    # 'Flood_Class': 'flood_class',
    # 'Erosion_Class': 'erosion_class',
}


VILLAGES_OF_ALL_THE_DISTRICTS = {
    'District_Name': 'district_name',
    'Revenue_circle': 'revenue_circle',
    'Village_name': 'village_name',
    'District_Code': 'district_code',
    'Village_Code': 'village_code',
    'Circle_Name': 'circle_name',
    'Block_Name': 'block_name',
    'Distance_from_Headquarter': 'distance_from_headquarter',
    'Total_Area': 'total_area',
    'Average_Elevation': 'average_elevation',
    'Topography': 'topography',

}

VILLAGE_ROAD_INFO_MAPPING = {
    'District_Name': 'district_name',
    'Village_Name': 'village_name',
    'Village_Id': 'village_code',  # this links to tblVillage via village_code
    'District_Code': 'district_code',
    'Road_Surface_Type': 'road_surface_type',
    'Road_Constructed_By': 'road_constructed_by',
    'Road_Length_(m)': 'road_length_m',
    'Flood_Depth(m)': 'flood_depth_m',
    'Flood_Class': 'flood_class'
}

VILLAGE_ROAD_INFO_EROSION = {
    'District_Name': 'district_name',
    'Village_Name': 'village_name',
    'District_Code': 'district_code',
    'Village_Id': 'village_code',  # links to tblVillage via village_code
    'Road_Surface_Type': 'road_surface_type',
    'Road_Constructed_By': 'road_constructed_by',
    'Road_Length_(m)': 'road_length_m',
    'Erosion_Class': 'erosion_class'
}

TRAINING_MASTER_LIST={
    'Activity':'name',
    'Activity_bn':"name_bn",
    "Activity_as":"name_as"
}

RESCUE_EQUEP_MASTER_LIST={
    'Task_Force':'name',
    'Task_Force_bn':'name_bn',
    'Task_Force_as':'name_as',
    'Item':"task_force",
    'Item_bn':"task_force_bn",
    'Item_as':"task_force_as",
    'Specification':'specification',
    'Specification_bn':'specification_bn',
    'Specification_as':'specification_as'
}

VDMP_ACTIVITIES={
    'Activity_Name':'name',
    'Activity_Name_bn':'name_bn',
    'Activity_Name_as':'name_as',
    'Sr. No':'order'
}


BRIDGE_SURVEY_INFO = {
    'Username': 'username',
    'spatial_id': 'spatial_id',
    'spatial_ref': 'spatial_ref',
    'polygon_id': 'polygon_id',
    'village_id': 'village_code',
    'village_name': 'village_name',
    'district_name': 'district_name',
    'survey_id': 'survey_id',
    'geometry': 'geometry',
    'user_id': 'user_id',
    'under_id': 'under_id',
    'unique_id': 'unique_id',
    'date': 'date',
    'form_id': 'form_id',
    'tab_id': 'tab_id',
    'tab_name': 'tab_name',

    'Bridge surface type': 'bridge_surface_type',
    'Length (meters)': 'length_meters',
    'Width(meters)': 'width_meters',
    'Photographs': 'photographs',
    'Bridge pillar material': 'bridge_pillar_material',
    'Number of pillars bridge has': 'number_of_pillars',
    'Deck material': 'deck_material',
    'Condition of deck': 'condition_of_deck',
    'General condition of bridge': 'general_condition',
    'Status of the access part of bridge': 'status_access_part',
    'Any other remarks': 'remarks',
}


RISK_ASSESMENT_MAPPING = {
    'Village_Name': 'village_name',
    'Vill_ID': 'village_code',
    'Hazard': 'hazard',
    'Exposure_Type': 'exposure_type',
    'Total Exposure Value (INR Crore)': 'total_exposure_value_inr_crore',
    'Loss (INR Crore)': 'loss_inr_crore',
    'Loss % wrt exposure value': 'loss_percent_wrt_exposure_value',
}

PRA_MAIN_MAPPING = {
    'Vill_ID': 'village_code',
    'Children Below 6 Years(Male)': 'children_below_6_male',
    'Children Below 6 Years(Female)': 'children_below_6_female',
    'Persons with Disability': 'persons_with_disability',
    'Persons with Chronic Illness': 'persons_with_chronic_illness',
    'Nearest PHC (km)': 'nearest_phc_km',
    'Nearest CHC (km)': 'nearest_chc_km',
    'Nearest hospital (km)': 'nearest_hospital_km',
    'Nearest veterinary  clinic (km)': 'nearest_veterinary_clinic_km',
    'Nearest Post Office (km)': 'nearest_post_office_km',
    'Nearest bank/ATM (km)': 'nearest_bank_atm_km',
    'Nearest Ambulance (km)': 'nearest_ambulance_km',
    'Nearest Bus service (km)': 'nearest_bus_service_km',
    'Main market (km)': 'main_market_km',
    'Nearest ration shop (km)': 'nearest_ration_shop_km',
    'Nearest High School (km)': 'nearest_high_school_km',
    'Nearest Higher secondary': 'nearest_higher_secondary_km',
    'Nearest College (km)': 'nearest_college_km',
    'Nearest Police station (km)': 'nearest_police_station_km',
    'Farmers availing  agriculture insurance': 'farmers_agriculture_insurance',
    'Farmers availing livestock insurance': 'farmers_livestock_insurance',
    'Occupational category': 'occupational_category',
    'Flood frequency': 'flood_frequency',
    'Flood severity': 'flood_severity',
    'Erosion hazard frequency': 'erosion_hazard_frequency',
    'Erosion hazard severity': 'erosion_hazard_severity',
    'Strong wind hazard frequency': 'strong_wind_hazard_frequency',
    'Strong wind hazard severity': 'strong_wind_hazard_severity',
    'Earthquake hazard frequency': 'earthquake_hazard_frequency',
    'Earthquake hazard severity': 'earthquake_hazard_severity',
    'Distance from District headquarter': 'distance_from_district_headquarter_km',
    'Average elevation (about MSL)': 'average_elevation_msl',
    'Farmers groups (Name and phone number)': 'farmers_groups',
    'Weavers groups (Name and phone number)': 'weavers_groups',
    'NGOs/CGOs (Name and phone number)': 'ngos_cgos',
    'Domestic solid waste collection and management (details of agency doing)': 'domestic_solid_waste_management',
    'Government schemes and person contact number': 'government_schemes_contact',
    'Siltation': 'siltation',
    'Water logging in agriculture land': 'water_logging_agriculture_land',
    'Encroachment of wetlands (paddy field and ponds/lakes) ': 'encroachment_of_wetlands',
    'Modification of natural drains': 'modification_of_natural_drains',
}

PRA_ASSETS_MAPPING = {
    'Village Name': 'village_name',
    'Vill_ID': 'vill_id',
    'Equipment': 'equipment',
    'Name of the owner': 'name_of_the_owner',
    'Phone Number': 'phone_number',
    'Remark': 'remark',
    'Asset Count': 'asset_count'
}

PRA_SHELTER_MAPPING = {
    'Village Name': 'village_name',
    'Vill_ID': 'vill_id',
    'Name of shelter': 'name_of_shelter',
    'Contact Person': 'contact_person',
    'Phone Number': 'phone_number',
    'Number of rooms': 'number_of_rooms',
    'Capacity': 'capacity',
    'Toilet facilities available': 'toilet_facility_available',
    'Drinking water available': 'drinking_water_facility_available',
    'Alternate power supply available': 'alternate_power_source'
}

LINE_DEPARTMENT_MAPPING = {
    'Vill_Id': 'vill_id',
    'Section': 'section',
    'Contact Name': 'contact_name',
    'Phone Number': 'phone_number',
    'Official':'official_number'
}

FGD_WASH_SUMMARY_MAPPING = {
    'Village Name': 'village_name',
    'District': 'district',
    'Dist_ID': 'dist_id',
    'Vill_ID': 'vill_id',
    'Drinking Water – Sources & Access': 'drinking_water_sources_access',
    'Adequacy & Reliability': 'adequacy_reliability',
    'Equity & Inclusion': 'equity_inclusion',
    'Affordability': 'affordability',
    'Water Quality': 'water_quality',
    'Traditional Practices': 'traditional_practices',
    'Community Role & JJM Implementation': 'community_role_jjm_implementation',
    'Infrastructure & Damage': 'infrastructure_damage',
    'Sanitation – Existing Facilities': 'sanitation_existing_facilities',
    'Impact of Floods': 'impact_of_floods',
    'Erosion Impact': 'erosion_impact',
    'Hygiene Practices': 'hygiene_practices',
    'Health Concerns': 'health_concerns',
    'Community Awareness': 'community_awareness',
    'Community Participation & Resilience': 'community_participation_resilience'
}

FGD_LIVELIHOOD_SUMMARY_MAPPING = {
    'Village Name': 'village_name',
    'District': 'district',
    'Dist_ID': 'dist_id',
    'Vill_ID': 'vill_id',
    'Cropping Pattern': 'cropping_pattern',
    'Cropping Calendar': 'cropping_calendar',
    'Livestock and Allied Activities': 'livestock_and_allied_activities',
    'Departmental Support': 'departmental_support',
    'Challenges in Agriculture': 'challenges_in_agriculture',
    'Challenges in Livestock': 'challenges_in_livestock'
}



import requests

def get_lat_lon(name):
    API_KEY = "pk.0e49d17193d1898c2e3082b3c7143384"
    place = f"{name}, Assam, India"
    url = f"https://us1.locationiq.com/v1/search?key={API_KEY}&q={place}&format=json"

    try:
        response = requests.get(url, timeout=5)
        if response.status_code == 200:
            data = response.json()
            if data:
                lat = data[0]['lat']
                lon = data[0]['lon']
                return lat, lon
        print(f"Failed to fetch location for: {name}")
    except Exception as e:
        print(f"Error: {e} for place: {name}")
    return None, None


def check_existing_village_data(df, model_class):
    """Check which villages already have data in the database"""
    village_codes = set()
    
    for _, row in df.iterrows():
        for col in ['vill_id', 'vill_id', 'village_code']:
            if col.lower() in [c.lower() for c in df.columns]:
                vill_code = str(row.get(col, '')).strip()
                if vill_code and vill_code.lower() != 'nan':
                    village_codes.add(vill_code)
                    break
    
    existing_villages = []
    for vill_code in village_codes:
        try:
            village = tblVillage.objects.get(code=vill_code)
            if model_class.objects.filter(village=village).exists():
                existing_villages.append(vill_code)
        except:
            continue
    
    return existing_villages
    # Update Districts


def store_lat_lon():
    all_districts = tblDistrict.objects.all()
    for district in all_districts:
        district_name = district.name.strip().title()
        lat, lon = get_lat_lon(district_name)

        if lat and lon:
            district.latitude = lat
            district.longitude = lon
        district.name = district_name
        district.save()
        print(f"District: {district.name}, Latitude: {lat}, Longitude: {lon}")

    # Update Circles
    all_circles = tblCircle.objects.all()
    for circle in all_circles:
        circle.name = circle.name.strip().title()
        circle.save()

    # Update Gram Panchayats
    all_grampanchayat = tblGramPanchayat.objects.all()
    for gp in all_grampanchayat:
        gp.name = gp.name.strip().title()
        gp.save()

    # Update Villages (once per district)
    all_villages = tblVillage.objects.select_related(
        "gram_panchayat", "gram_panchayat__circle", "gram_panchayat__circle__district"
    )

    for village in all_villages:
        village_name = f"{village.name}, {village.gram_panchayat.circle.district.name}"
        lat, lon = get_lat_lon(village_name)

        if lat and lon:
            village.latitude = lat
            village.longitude = lon

        village.name = village.name.strip().title()
        village.save()
        print(f"Village: {village.name}, Latitude: {lat}, Longitude: {lon}")
