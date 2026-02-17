import argparse
import os
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook
import csv
import psycopg2


DEFAULT_SHEET = "all themes intervention"
DEFAULT_EXCEL = r"D:\AssamCRVV2\portalcode\assamcrv\assam_crv\file\mitigation_intervention_master_table_cleaned.csv"
DEFAULT_DB = "assam_crv2"
DEFAULT_TABLE = "mitigation_mitigationinterventionmaster"


def load_env_file(path):
    data = {}
    if not path or not os.path.exists(path):
        return data
    with open(path, "r", encoding="utf-8") as handle:
        for raw in handle:
            line = raw.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            value = value.strip().strip('"').strip("'")
            data[key.strip()] = value
    return data


def clean_text(value):
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() == "null":
        return ""
    return text


def clean_decimal(value):
    if value is None:
        return None
    if isinstance(value, str) and value.strip().lower() == "null":
        return None
    text = str(value).replace(",", "").strip()
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def get_db_config(args):
    env = dict(os.environ)
    env.update(load_env_file(args.env_file))
    return {
        "dbname": args.dbname or env.get("DB_NAME") or "assamcrv_portal_v2",
        "user": args.user or env.get("DB_USER") or "postgres",
        "password": args.password or env.get("DB_PASSWORD") or "pass@123",
        "host": args.host or env.get("DB_HOST") or "localhost",
        "port": int(args.port or env.get("DB_PORT") or 5434),
    }


def ensure_table(conn, table):
    with conn.cursor() as cur:
        cur.execute(
            f"""
            CREATE TABLE IF NOT EXISTS {table} (
                id BIGSERIAL PRIMARY KEY,
                theme VARCHAR(200) NOT NULL,
                subtheme VARCHAR(200) NOT NULL,
                vulnerable_asset VARCHAR(200),
                intervention_type VARCHAR(200),
                intervention_name TEXT NOT NULL,
                display_note TEXT,
                unit VARCHAR(50),
                default_quantity NUMERIC(12, 2),
                unit_cost_rs NUMERIC(12, 2),
                status VARCHAR(50) NOT NULL DEFAULT 'active'
            );
            """
        )


def iter_excel_rows(excel_path, sheet_name):
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name}")
    ws = wb[sheet_name]
    return ws.iter_rows(values_only=True)


def iter_csv_rows(csv_path):
    def normalize_header(header):
        return header.strip() if isinstance(header, str) else header

    with open(csv_path, newline="", encoding="utf-8") as handle:
        reader = csv.reader(handle)
        for row in reader:
            yield [normalize_header(cell) for cell in row]


def import_data(conn, source_path, sheet_name, table, truncate):
    ext = os.path.splitext(source_path)[1].lower()
    if ext == ".csv":
        rows = iter_csv_rows(source_path)
    else:
        rows = iter_excel_rows(source_path, sheet_name)

    headers = next(rows, None)
    if not headers:
        return 0

    header_map = {str(h).strip(): idx for idx, h in enumerate(headers) if h is not None}

    def get_cell(row, key):
        idx = header_map.get(key)
        return row[idx] if idx is not None and idx < len(row) else None

    required = ["Theme", "Sub theme", "Mitigation intervention"]
    for req in required:
        if req not in header_map:
            raise ValueError(f"Missing required column: {req}")

    insert_sql = f"""
        INSERT INTO {table} (
            theme,
            subtheme,
            vulnerable_asset,
            intervention_type,
            intervention_name,
            display_note,
            unit,
            default_quantity,
            unit_cost_rs,
            status
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """

    count = 0
    with conn.cursor() as cur:
        if truncate:
            cur.execute(f"TRUNCATE TABLE {table} RESTART IDENTITY;")

        for row in rows:
            theme = clean_text(get_cell(row, "Theme"))
            subtheme = clean_text(get_cell(row, "Sub theme"))
            intervention_name = clean_text(get_cell(row, "Mitigation intervention"))
            if not theme or not subtheme or not intervention_name:
                continue

            data = (
                theme,
                subtheme,
                clean_text(get_cell(row, "Vulnerable asset")),
                clean_text(get_cell(row, "Intervention type")),
                intervention_name,
                clean_text(
                    get_cell(
                        row,
                        "Display as Note: Explanation of mitigation intervention/Guiding points for repair",
                    )
                ),
                clean_text(get_cell(row, "Unit")),
                clean_decimal(get_cell(row, "Quantity")),
                clean_decimal(get_cell(row, "Unit cost (Rs)")),
                "active",
            )

            cur.execute(insert_sql, data)
            count += 1

    return count


def main():
    parser = argparse.ArgumentParser(description="Create mitigation master table and import Excel data.")
    parser.add_argument("--excel", default=DEFAULT_EXCEL, help="Excel file path.")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="Sheet name to import.")
    parser.add_argument("--table", default=DEFAULT_TABLE, help="Target table name.")
    parser.add_argument("--dbname", default=None, help="Database name (default: assam_crv2).")
    parser.add_argument("--host", default=None, help="Database host.")
    parser.add_argument("--port", default=None, help="Database port.")
    parser.add_argument("--user", default=None, help="Database user.")
    parser.add_argument("--password", default=None, help="Database password.")
    parser.add_argument("--env-file", default=r"E:\Siraj\assam_crv\assam_crv\.env", help="Path to .env file.")
    parser.add_argument("--truncate", action="store_true", help="Truncate table before import.")

    args = parser.parse_args()

    if not os.path.exists(args.excel):
        raise SystemExit(f"Excel file not found: {args.excel}")

    db_config = get_db_config(args)
    conn = psycopg2.connect(**db_config)
    try:
        ensure_table(conn, args.table)
        inserted = import_data(conn, args.excel, args.sheet, args.table, args.truncate)
        conn.commit()
        print(f"Inserted {inserted} rows into {args.table}.")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
